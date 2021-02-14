import os
from openpyxl import load_workbook


# TEMP_DOCUMENT_PATH = "temp/person_doc.xlsm"
# TEMPLATE_PATH = "data/print/template.xlsm"

def export_record_to_doc(data):
    """ Create document to print according to data 'data' """

    document_path = "data/print/document.xlsm"
    wb = load_workbook(document_path)
    sheet = wb["tlac"]

    sheet.cell(row=6, column=6, value=data["fname"])
    sheet.cell(row=6, column=7, value=data["lname"])
    sheet.cell(row=8, column=6, value=data["residence"])
    sheet.cell(row=10, column=6, value=data["birth_date"])
    sheet.cell(row=18, column=4, value=data["today"])

    wb.save(document_path)

    os.system("xdg-open data/print/document.xlsm")
